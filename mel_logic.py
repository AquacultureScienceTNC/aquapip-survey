"""
AquaPIP — MEL Monitoring Survey
Core data + matching logic.

This module is deliberately independent of Streamlit so it can be unit-tested
on its own. app.py imports from here.

KEY DESIGN CHOICES (read before editing)
----------------------------------------
1.  The database is read from data/mel_database.xlsx by COLUMN HEADER NAME,
    never by fixed position. That means:
      * When you add columns AK ("Protocol Method Summary") and later
        AM/AN/AO ("Protocol Template URL" / "Data Sheet URL" /
        "Stats Workbook URL"), they are picked up automatically. Their physical
        column letter does not matter — only the header text.
      * Re-ordering columns will not break the app.
      * Missing columns simply read as empty (so v22 works today, v23+ works
        tomorrow with no code change).

2.  Survey option lists (indicators + farm-profile fields) are built from the
    UNION of the workbook's "Vocabularies" sheet and the distinct values that
    actually appear in the data. This is what makes "change a tag -> survey
    changes" true, AND it keeps emerging indicators reachable (e.g. CC 3.1.1
    Wave attenuation, WQ 1.1.2 Denitrification, WQ 1.3.1 Water filtration exist
    in the data but are not yet in the controlled vocabulary).

3.  Indicator matching is done on the indicator CODE (e.g. "WQ 1.1.1"), not the
    full label. Many rows carry annotated labels like
    "WQ 1.1.1 — Bioextraction (tissue C/N + isotopes)". Matching on the code
    collapses those variants onto the canonical indicator the farmer selected.

4.  Multi-select cells are split on TOP-LEVEL commas only. Several cells contain
    commas inside parentheses, e.g. "WQ 1.2.1 — Inorganic chemistry (DO,
    salinity)". A naive comma split would shatter those into junk tokens.
"""

import os
import re
from collections import OrderedDict

import openpyxl

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

DATA_SHEET_CANDIDATES = ["in"]          # preferred data sheet name(s)
VOCAB_SHEET = "Vocabularies"

# Tier -> label + colour (your palette from the Usability Tier System).
TIER_META = OrderedDict([
    (4, {"label": "Practitioner", "color": "#92D050", "skill": "Basic"}),
    (3, {"label": "Partnership",  "color": "#C5E0B4", "skill": "Intermediate"}),
    (2, {"label": "Researcher",   "color": "#FFD966", "skill": "Advanced"}),
    (1, {"label": "Reference",    "color": "#F08080", "skill": "Reference"}),
])
UNRATED = {"label": "Unrated", "color": "#CCCCCC", "skill": "—"}

# Indicator code pattern, e.g. "H&B 1.1.1", "WQ 1.1.1", "CC 3.1.1".
CODE_RE = re.compile(r'(H&B|WQ|CC)\s+\d+\.\d+\.\d+')

# Values that mean "applies broadly" — they should not penalise a protocol in
# farm-profile matching (a protocol tagged "Not species-specific" fits any farm).
UNIVERSAL_VALUES = {
    "not species-specific", "not algae-specific", "not gear-specific",
    "not scale-specific", "not depth-specific", "not exposure-specific",
    "not site-specific", "not specific", "not applicable", "not gear specific",
    "not scale specific", "not depth specific", "not exposure specific",
    "not time-specific", "cross-zone / global", "not scale-specific",
}

# Farm-profile fields: (internal key, display label, Vocabularies header, record key)
PROFILE_FIELDS = [
    ("aqua_type",      "Aquaculture type",   "Aquaculture Type",   "aqua_type"),
    ("species",        "Cultivated species", "Cultivated Species", "species"),
    ("algae_type",     "Algae type",         "Algae Type",         "algae_type"),
    ("farm_structure", "Farm structure",     "Farm Structure",     "farm_structure"),
    ("farm_scale",     "Farm scale",         "Farm Scale",         "farm_scale"),
    ("coculture",      "Co-culture context", "Co-culture Context", "coculture"),
    ("climate_zone",   "Climate zone",       "Climate Zone",       "climate_zone"),
    ("site_depth",     "Site depth",         "Site Depth",         "site_depth"),
    ("wave_energy",    "Wave energy",         "Wave Energy",        "wave_energy"),
    ("water_clarity",  "Water clarity",      "Water Clarity",      "water_clarity"),
]

# Fields used for the soft "matches your farm" score. is_list flags multi-select.
_PROFILE_MATCH = [
    ("aqua_type", True), ("species", True), ("algae_type", True),
    ("farm_structure", False), ("farm_scale", False), ("coculture", False),
    ("climate_zone", False), ("site_depth", False), ("wave_energy", False),
    ("water_clarity", False),
]

# ------------------------------------------------------------------ #
# FARMER-FACING GOAL SCREEN  (the curated "option C" layer)
# ------------------------------------------------------------------ #
# This is the ONLY hand-curated mapping in the tool. Page 1 asks a practitioner
# what they want to measure in plain terms; that pre-selects the MEL indicator
# codes below (which the practitioner can then trim or extend on the indicator
# step). Every code that is NOT attached to a goal still appears in the manual
# indicator picker, so nothing in the database is unreachable.
#
# TO EDIT: add/rename a goal, or change its list of indicator codes. Codes must
# match the code portion of an indicator that exists in the vocab or data
# (e.g. "WQ 1.1.1"). Codes with no matching rows simply show "no protocol yet".
FARMER_GOALS = OrderedDict([
    ("Blue Carbon — ecosystem-positive farming", {
        "blurb": ("Enhancing blue-carbon ecosystems via filtration, nutrient & "
                  "turbidity reduction, and habitat protection (Alleway et al. "
                  "2025). This is an ecosystem-enhancement framing, not a "
                  "carbon-sink / sequestration claim."),
        "codes": ["WQ 1.1.1", "WQ 1.1.2", "WQ 1.3.1", "WQ 2.1.1",
                  "H&B 1.4.1", "CC 1.2.1", "CC 2.1.1", "CC 3.1.1"],
    }),
    ("Fisheries support & benefits", {
        "blurb": ("Habitat provision and biodiversity that support wild fish and "
                  "mobile fauna around the farm."),
        "codes": ["H&B 1.1.1", "H&B 1.2.1", "H&B 2.1.1", "H&B 2.2.1"],
    }),
    ("Water Quality", {
        "blurb": ("Nutrient extraction, dissolved oxygen, chlorophyll-a and "
                  "plankton condition."),
        "codes": ["WQ 1.1.1", "WQ 1.1.2", "WQ 1.3.1", "WQ 1.2.1",
                  "WQ 2.1.1", "WQ 2.1.2"],
    }),
    ("Benthic Health", {
        "blurb": "Condition of the seabed beneath and around the farm.",
        "codes": ["H&B 1.3.1", "H&B 1.4.1"],
    }),
])


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #

def _s(v):
    """None-safe stripped string."""
    return "" if v is None else str(v).strip()


def _norm(s):
    """Normalise a header/value for tolerant lookups (dash + case + spacing)."""
    return re.sub(r"\s+", " ",
                  str(s).lower().replace("–", "-").replace("—", "-")).strip()


def split_multi(value):
    """Split a multi-select cell on top-level commas only (ignore commas that
    sit inside (), [] or {})."""
    if value is None:
        return []
    out, buf, depth = [], [], 0
    for ch in str(value):
        if ch in "([{":
            depth += 1
            buf.append(ch)
        elif ch in ")]}":
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch == "," and depth == 0:
            tok = "".join(buf).strip()
            if tok:
                out.append(tok)
            buf = []
        else:
            buf.append(ch)
    tok = "".join(buf).strip()
    if tok:
        out.append(tok)
    return out


def indicator_code(token):
    """Reduce an indicator label to its matching key (the code)."""
    if not token:
        return None
    m = CODE_RE.search(token)
    if m:
        return re.sub(r"\s+", " ", m.group(0)).strip()
    t = token.lower()
    if "cross-cutting" in t:
        return "Cross-cutting"
    if "not in mel" in t:
        return "Not in MEL framework"
    return re.sub(r"\s+", " ", token).strip()


def goal_of_code(code):
    if not code:
        return None
    if code.startswith("H&B"):
        return "Habitat & Biodiversity"
    if code.startswith("WQ"):
        return "Water Quality"
    if code.startswith("CC"):
        return "Climate Change"
    if code == "Cross-cutting":
        return "Cross-cutting / Framework"
    return None


def clean_label(token):
    """Display label: drop parenthetical annotations, tidy whitespace."""
    lab = re.sub(r"\s*\([^)]*\)", "", token).strip()
    return re.sub(r"\s+", " ", lab)


def _code_sort_key(code):
    m = re.search(r"(\d+)\.(\d+)\.(\d+)", code)
    return tuple(int(x) for x in m.groups()) if m else (99, 99, 99)


def parse_tier(value):
    """Return (n, label, color, skill)."""
    s = _s(value)
    m = re.match(r"\s*([1-4])", s)
    if m:
        n = int(m.group(1))
        meta = TIER_META[n]
        return n, meta["label"], meta["color"], meta["skill"]
    low = s.lower()
    for n, meta in TIER_META.items():
        if meta["label"].lower() in low:
            return n, meta["label"], meta["color"], meta["skill"]
    return 0, UNRATED["label"], UNRATED["color"], UNRATED["skill"]


def cost_bucket(value):
    """Return (label, ordinal 0..3). 0 == n/a or unknown.

    Ranges resolve to their UPPER bound (ceiling), so a farmer-facing budget
    isn't understated: 'Low–Medium' -> Medium, 'Medium–High' -> High. We read
    the rating head (text before an em-dash detail clause) and take the max of
    the standalone level words found there, skipping compounds like
    'high-resolution' / 'low-cost' / 'medium-term'."""
    s = _s(value)
    low = s.lower()
    if not s or low.startswith("n/a") or low.startswith("na "):
        return ("n/a", 0)
    head = re.split(r"\s+—\s+", s, 1)[0]   # drop trailing narrative ("... sieves Low")
    order = {"low": 1, "medium": 2, "med": 2, "high": 3}
    found = []
    for m in re.finditer(r"(?i)\b(low|medium|med|high)\b", head):
        if re.match(r"-[A-Za-z]", head[m.end():m.end() + 2]):
            continue   # skip compound words (high-resolution, low-cost, medium-term)
        found.append(order[m.group(1).lower()])
    if found:
        c = max(found)
        return ({1: "Low", 2: "Medium", 3: "High"}[c], c)
    return ("—", 0)


# Controlled Sampling Effort phrase -> effort ordinal (Sub-rule 6a).
# Cells are written as "<phrase> — optional detail"; the leading phrase sets the level.
EFFORT_LEVEL = {
    "Single deployment (hours to 1 day)": 1,
    "Short campaign (days to 2 weeks)": 1,
    "Seasonal (1–3 months)": 2,
    "Multi-season (3–9 months)": 2,
    "Year-long study (12 months)": 3,
    "Multi-year monitoring (> 1 year)": 3,
    "Continuous (sensor / autonomous)": 3,
    "One-time (no sampling — desk-based)": 1,
    "Not time-specific": 0,
}
_EFFORT_LABEL = {0: "—", 1: "Low", 2: "Medium", 3: "High"}


def effort_bucket(value):
    """Return (label, ordinal 0..3) from the sampling-effort value.

    Prefers an exact controlled-vocab leading phrase (Sub-rule 6a format:
    '<phrase> — detail'); falls back to keyword heuristics for any
    non-conformant stragglers, so behaviour is unchanged pre-standardisation
    and exact once the column is standardised."""
    s = _s(value)
    if not s:
        return ("—", 0)
    for phrase in sorted(EFFORT_LEVEL, key=len, reverse=True):
        if s.startswith(phrase):
            o = EFFORT_LEVEL[phrase]
            return (_EFFORT_LABEL[o], o)
    # ---- fallback heuristic (unrecognised / pre-standardisation values) ----
    low = s.lower()
    if "continuous" in low:
        return ("High", 3)
    high_kw = ["multi-year", "multi year", "multi-decade", "year-long",
               "12-month", "12 month", "annual", "quarterly", "8+", "5-year",
               "5 yr", "5-yr", "hindcast", "> 1 year", ">1 year"]
    if any(k in low for k in high_kw):
        return ("High", 3)
    if "multi-season" in low or "multi season" in low or "seasonal" in low:
        return ("Medium", 2)
    if "short campaign" in low or "single deployment" in low:
        return ("Low", 1)
    if any(k in low for k in ["one-time", "one time", "desk", "model", "discrete"]):
        return ("Low", 1)
    if low.startswith("not "):
        return ("—", 0)
    return ("Medium", 2)


def dots(ordinal):
    """Cost/effort glyph: low=1 -> ●○○, medium=2 -> ●●○, high=3 -> ●●●."""
    if ordinal <= 0:
        return "—"
    ordinal = max(1, min(3, ordinal))
    return "●" * ordinal + "○" * (3 - ordinal)


def _is_universal(v):
    return _norm(v) in UNIVERSAL_VALUES


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #

def load_database(path):
    """Read the workbook. Returns (rows, vocab, headers).

    rows   : list of dicts, one per protocol/reference.
    vocab  : dict header -> list of controlled values (from Vocabularies sheet).
    headers: dict of the data sheet's header -> column index (for reference).
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    data_ws = None
    for name in DATA_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            data_ws = wb[name]
            break
    if data_ws is None:
        for name in wb.sheetnames:
            if name != VOCAB_SHEET:
                data_ws = wb[name]
                break

    # header maps (exact + normalised)
    headers, headers_norm = {}, {}
    for c in range(1, data_ws.max_column + 1):
        h = data_ws.cell(row=1, column=c).value
        if h is not None and str(h).strip():
            headers[str(h).strip()] = c
            headers_norm[_norm(h)] = c

    def cell(r, header):
        c = headers.get(header)
        if c is None:
            c = headers_norm.get(_norm(header))
        return data_ws.cell(row=r, column=c).value if c else None

    rows = []
    for r in range(2, data_ws.max_row + 1):
        title = cell(r, "Title")
        authors = cell(r, "Authors")
        if title is None and authors is None:
            continue  # skip blank rows

        ind_tokens = split_multi(cell(r, "MEL Indicator (multi-select)"))
        codes = sorted({indicator_code(t) for t in ind_tokens if indicator_code(t)},
                       key=_code_sort_key)
        tier_n, tier_label, tier_color, skill = parse_tier(cell(r, "Usability Rating"))
        cost_label, cost_ord = cost_bucket(cell(r, "Estimated Cost"))
        eff_label, eff_ord = effort_bucket(cell(r, "Sampling Effort (deployment time)"))

        rows.append({
            "row": r,
            "authors": _s(authors),
            "title": _s(title),
            "publication": _s(cell(r, "Publication")),
            "year": _s(cell(r, "Year")),
            "indicators_matched": _s(cell(r, "MEL Indicator(s) Matched")),
            "protocol_or_reference": _s(cell(r, "Protocol or Reference?")),
            "is_protocol": _s(cell(r, "Protocol or Reference?")).lower().startswith("protocol"),
            "measures": _s(cell(r, "Adequately Measures Indicator? (1–2 sentences)")),
            "exec_practitioner": _s(cell(r, "Executable by Practitioner (Farmer)?")),
            "exec_researcher": _s(cell(r, "Executable by Researcher / Academic?")),
            "equipment": _s(cell(r, "Equipment Type")),
            "cost_text": _s(cell(r, "Estimated Cost")),
            "cost_label": cost_label, "cost_ord": cost_ord,
            "effort_text": _s(cell(r, "Sampling Effort (deployment time)")),
            "effort_label": eff_label, "effort_ord": eff_ord,
            "stats": _s(cell(r, "Statistical Approach")),
            "openly_accessible": _s(cell(r, "Openly Accessible?")),
            "source_read": _s(cell(r, "Source Read")),
            "url": _s(cell(r, "Access URL / Repository")),
            "notes": _s(cell(r, "Notes / Caveats")),
            "tier": tier_n, "tier_label": tier_label,
            "tier_color": tier_color, "skill": skill,
            # tag columns
            "aqua_type": split_multi(cell(r, "Aquaculture Type (multi-select)")),
            "mel_goal": split_multi(cell(r, "MEL Goal (multi-select)")),
            "_indicator_tokens": ind_tokens,
            "indicator_codes": codes,
            "climate_zone": _s(cell(r, "Climate Zone")),
            "species": split_multi(cell(r, "Cultivated Species (multi-select)")),
            "farm_structure": _s(cell(r, "Farm Structure")),
            "farm_scale": _s(cell(r, "Farm Scale")),
            "water_clarity": _s(cell(r, "Water Clarity")),
            "coculture": _s(cell(r, "Co-culture Context")),
            "method_category": split_multi(cell(r, "Method Category (multi-select)")),
            "algae_type": split_multi(cell(r, "Algae Type (multi-select)")),
            "site_depth": _s(cell(r, "Site Depth")),
            "wave_energy": _s(cell(r, "Wave Energy")),
            # ---- future columns (read by name; empty until you add them) ----
            "method_summary": _s(cell(r, "Protocol Method Summary")),   # AK
            "url_template": _s(cell(r, "Protocol Template URL")),        # AM
            "url_datasheet": _s(cell(r, "Data Sheet URL")),             # AN
            "url_workbook": _s(cell(r, "Stats Workbook URL")),          # AO
        })

    return rows, load_vocab(wb), headers


def load_vocab(wb):
    """Read the Vocabularies sheet -> {header: [values]}. Header row is detected
    (the row containing 'Aquaculture Type')."""
    if VOCAB_SHEET not in wb.sheetnames:
        return {}
    vs = wb[VOCAB_SHEET]
    header_row = 6
    for r in range(1, min(vs.max_row, 15) + 1):
        rowvals = [str(vs.cell(row=r, column=c).value or "")
                   for c in range(1, vs.max_column + 1)]
        if any("Aquaculture Type" in v for v in rowvals):
            header_row = r
            break
    cols = {}
    for c in range(1, vs.max_column + 1):
        h = vs.cell(row=header_row, column=c).value
        if not h:
            continue
        vals = []
        for r in range(header_row + 1, vs.max_row + 1):
            v = vs.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                vals.append(str(v).strip())
        cols[str(h).strip()] = vals
    return cols


# --------------------------------------------------------------------------- #
# Building survey option lists
# --------------------------------------------------------------------------- #

def build_indicator_options(vocab_indicators, rows):
    """Grouped indicator options for the survey.

    Returns OrderedDict: MEL goal -> [ {code, label}, ... ].
    Union of vocab + codes present in data. Excludes the meta tags
    'Cross-cutting' and 'Not in MEL framework' (they are not measurable
    indicators a practitioner selects)."""
    by_code = OrderedDict()  # code -> display label (vocab label preferred)
    for lab in vocab_indicators:
        code = indicator_code(lab)
        if code and code.startswith(("H&B", "WQ", "CC")):
            by_code.setdefault(code, clean_label(lab))
    for row in rows:
        for tok in row["_indicator_tokens"]:
            code = indicator_code(tok)
            if code and code.startswith(("H&B", "WQ", "CC")) and code not in by_code:
                by_code[code] = clean_label(tok)

    groups = OrderedDict([
        ("Habitat & Biodiversity", []),
        ("Water Quality", []),
        ("Climate Change", []),
    ])
    for code, label in by_code.items():
        g = goal_of_code(code)
        if g in groups:
            groups[g].append({"code": code, "label": label})
    for g in groups:
        groups[g].sort(key=lambda d: _code_sort_key(d["code"]))
    return groups


def profile_options(vocab, rows):
    """For each farm-profile field: vocab values first, then any extra data
    values, de-duplicated and order-preserving."""
    opts = {}
    for fkey, _label, vheader, rkey in PROFILE_FIELDS:
        seen = []
        for v in vocab.get(vheader, []):
            if v not in seen:
                seen.append(v)
        for row in rows:
            val = row[rkey]
            vals = val if isinstance(val, list) else ([val] if val else [])
            for x in vals:
                if x and x not in seen:
                    seen.append(x)
        opts[fkey] = seen
    return opts


def flatten_options(ind_groups):
    """Maps across all indicator options: code<->label and code->goal."""
    code2label, label2code, code2goal = {}, {}, {}
    for goal, items in ind_groups.items():
        for it in items:
            code2label[it["code"]] = it["label"]
            label2code[it["label"]] = it["code"]
            code2goal[it["code"]] = goal
    return code2label, label2code, code2goal


def codes_for_farmer_goals(farmer_goals, available_codes):
    """Union of indicator codes pre-selected by the chosen farmer goals,
    filtered to codes that actually exist as options."""
    out = []
    for g in farmer_goals:
        for c in FARMER_GOALS.get(g, {}).get("codes", []):
            if c in available_codes and c not in out:
                out.append(c)
    return out


# --------------------------------------------------------------------------- #
# Matching + ranking
# --------------------------------------------------------------------------- #

def profile_score(row, profile):
    """Soft score: +1 per specific tag match, +0.5 per 'applies broadly' tag.
    Returns (score, [matched field keys with a SPECIFIC match])."""
    score = 0.0
    matched = []
    for pkey, _is_list in _PROFILE_MATCH:
        user_val = profile.get(pkey)
        if not user_val:
            continue
        rowval = row.get(pkey)
        rowvals = rowval if isinstance(rowval, list) else ([rowval] if rowval else [])
        if not rowvals:
            continue
        if any(_is_universal(x) for x in rowvals):
            score += 0.5
            continue
        if any(_norm(x) == _norm(user_val) for x in rowvals):
            score += 1.0
            matched.append(pkey)
    return score, matched


def _aqua_match(farmer_aqua, row_aqua_list):
    """True if a protocol (with its aquaculture-type tags) applies to the
    farmer's sector. Aquaculture type is a HARD-ish filter (it decides which
    framework indicators apply); other profile fields stay soft."""
    fa = _norm(farmer_aqua)
    if not fa or fa == "not specific":
        return True
    rows = [_norm(x) for x in (row_aqua_list or [])]
    if not rows or any(r in ("not specific", "not applicable", "cross-system") for r in rows):
        return True  # universally applicable protocol
    if "multi-trophic" in fa or "imta" in fa:
        comps = {"seaweed", "mollusk", "echinoderm", "finfish", "multi-trophic (imta)"}
        return any(r in comps for r in rows)
    return fa in rows or any("multi-trophic" in r for r in rows)


def match_protocols(rows, selected_codes, profile, farmer_aqua=None):
    """For each selected indicator code, return ranked candidate protocols.

    Returns OrderedDict: code -> [ {row, pscore, pmatched, badge}, ... ]
    ranked best-first. If farmer_aqua is given, only protocols applicable to
    that sector are returned (resolves the CC sector-collision). 'badge' is
    True when the farm is a clear fit on the soft profile fields."""
    result = OrderedDict()
    for code in selected_codes:
        cands = []
        for row in rows:
            if code in row["indicator_codes"] and (
                    farmer_aqua is None or _aqua_match(farmer_aqua, row["aqua_type"])):
                pscore, pmatched = profile_score(row, profile)
                badge = len(pmatched) >= 2
                cands.append({"row": row, "pscore": pscore,
                              "pmatched": pmatched, "badge": badge})
        cands.sort(key=lambda d: (
            -d["row"]["tier"],                              # higher tier first
            0 if d["row"]["is_protocol"] else 1,            # protocol before reference
            -d["pscore"],                                   # better farm fit first
            d["row"]["cost_ord"] if d["row"]["cost_ord"] > 0 else 99,   # cheaper first
            d["row"]["effort_ord"] if d["row"]["effort_ord"] > 0 else 99,  # less effort first
            d["row"]["row"],                                # stable
        ))
        result[code] = cands
    return result


def goals_from_codes(codes):
    out = []
    for c in codes:
        g = goal_of_code(c)
        if g and g not in out:
            out.append(g)
    return out


# --------------------------------------------------------------------------- #
# MEL Indicator Reference table (sector-aware picker + education content)
# --------------------------------------------------------------------------- #

REF_INDICATOR_SHEET = "MEL_Indicators"
REF_SPECIES_SHEET = "Species_by_Sector"
REF_STRUCTURE_SHEET = "Structure_by_Sector"

# aquaculture-type (data vocab) -> reference-table sector(s)
_AQUA_SECTOR = {
    "seaweed": ["Seaweed"],
    "mollusk": ["Molluscs & Echinoderms"],
    "mollusc": ["Molluscs & Echinoderms"],
    "echinoderm": ["Molluscs & Echinoderms"],
    "finfish": ["Marine Finfish"],
    "marine finfish": ["Marine Finfish"],
}
_ALL_SECTORS = ["Seaweed", "Molluscs & Echinoderms", "Marine Finfish"]


def sectors_for_aqua(aqua_type):
    """Reference sector(s) that a farmer's aquaculture type maps to."""
    a = _norm(aqua_type)
    if "multi-trophic" in a or "imta" in a:
        return list(_ALL_SECTORS)
    if not a or a == "not specific":
        return list(_ALL_SECTORS)
    return _AQUA_SECTOR.get(a, list(_ALL_SECTORS))


def load_reference(path):
    """Read MEL_Indicator_Reference.xlsx -> {'indicators': [...], 'species': {sector:[...]}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    indicators = []
    if REF_INDICATOR_SHEET in wb.sheetnames:
        ws = wb[REF_INDICATOR_SHEET]
        H, Hn = {}, {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h:
                H[str(h).strip()] = c
                Hn[_norm(h)] = c

        def cell(r, name):
            col = H.get(name) or Hn.get(_norm(name))
            return _s(ws.cell(row=r, column=col).value) if col else ""

        for r in range(2, ws.max_row + 1):
            label = cell(r, "Indicator (label + code)") or cell(r, "Indicator")
            if not label:
                continue
            indicators.append({
                "sector": cell(r, "Sector"),
                "area": cell(r, "MEL Goal Area"),
                "label": label,
                "code": indicator_code(label),
                "goal": cell(r, "Goal"),
                "objective": cell(r, "Objective"),
                "measured": cell(r, "What is measured"),
                "complexity": cell(r, "Complexity"),
                "metric": cell(r, "Metric"),
                "method": cell(r, "Suggested method"),
                "proxy": cell(r, "Proxy / additional method"),
                "frequency": cell(r, "Frequency / Timing"),
                "location": cell(r, "Location of sampling"),
            })
    species = {}
    if REF_SPECIES_SHEET in wb.sheetnames:
        ws = wb[REF_SPECIES_SHEET]
        for r in range(2, ws.max_row + 1):
            sec = _s(ws.cell(row=r, column=1).value)
            sp = _s(ws.cell(row=r, column=2).value)
            if sec and sp:
                species.setdefault(sec, []).append(sp)
    structures = {}
    if REF_STRUCTURE_SHEET in wb.sheetnames:
        ws = wb[REF_STRUCTURE_SHEET]
        for r in range(2, ws.max_row + 1):
            sec = _s(ws.cell(row=r, column=1).value)
            stv = _s(ws.cell(row=r, column=2).value)
            if sec and stv:
                structures.setdefault(sec, []).append(stv)
    return {"indicators": indicators, "species": species, "structures": structures}


def indicators_for_sector(reference, aqua_type):
    """All reference indicators applicable to the farmer's aquaculture type,
    in framework order (sector, then goal area, then code)."""
    sectors = sectors_for_aqua(aqua_type)
    order_area = {"Habitat & Biodiversity": 0, "Water Quality": 1, "Climate Change": 2}
    out = [i for i in reference["indicators"] if i["sector"] in sectors]
    out.sort(key=lambda i: (sectors.index(i["sector"]) if i["sector"] in sectors else 9,
                            order_area.get(i["area"], 9), _code_sort_key(i["code"])))
    return out


def sector_indicator_groups(reference, aqua_type):
    """Grouped indicator options for the survey picker (by MEL goal area),
    filtered to the farmer's sector(s), with sector-correct labels.
    Returns OrderedDict area -> [ {code, label}, ... ]."""
    groups = OrderedDict([("Habitat & Biodiversity", []),
                          ("Water Quality", []),
                          ("Climate Change", [])])
    seen = set()
    for ind in indicators_for_sector(reference, aqua_type):
        if ind["label"] in seen:      # union dedupe (matters only for IMTA)
            continue
        seen.add(ind["label"])
        if ind["area"] in groups:
            groups[ind["area"]].append({"code": ind["code"], "label": ind["label"]})
    return groups


def species_options_for(reference, aqua_type):
    """Cultivated-species options for the chosen aquaculture type ('Any' rows
    always included; IMTA / Not-specific return the union)."""
    sp = reference.get("species", {})
    always = list(sp.get("Any", []))
    a = _norm(aqua_type)
    if not a or a == "not specific":
        allsp = [x for k, v in sp.items() if k != "Any" for x in v]
        return allsp + always
    if "multi-trophic" in a or "imta" in a:
        out = []
        for k in ("Seaweed", "Mollusk", "Finfish", "Echinoderm"):
            out += sp.get(k, [])
        return out + always
    # match the sheet's sector key case-insensitively
    for k, v in sp.items():
        if _norm(k) == a:
            return list(v) + always
    return always


def structure_options_for(reference, aqua_type):
    """Farm-structure options for the chosen aquaculture type ('Any' rows always
    included; IMTA / Not-specific return the union)."""
    stmap = reference.get("structures", {})
    always = list(stmap.get("Any", []))
    a = _norm(aqua_type)
    if not a or a == "not specific":
        allst = [x for k, v in stmap.items() if k != "Any" for x in v]
        return allst + always
    if "multi-trophic" in a or "imta" in a:
        out = []
        for k in ("Seaweed", "Mollusk", "Finfish", "Echinoderm"):
            out += stmap.get(k, [])
        return out + always
    for k, v in stmap.items():
        if _norm(k) == a:
            return list(v) + always
    return always


def algae_applicable(aqua_type):
    """Algae type only applies to seaweed and multi-trophic systems."""
    a = _norm(aqua_type)
    return ("seaweed" in a) or ("multi-trophic" in a) or ("imta" in a)


# --------------------------------------------------------------------------- #
# Self-test (run: python mel_logic.py)
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(here, "data", "mel_database.xlsx")
    rows, vocab, headers = load_database(db)
    print(f"Loaded {len(rows)} rows; {len(headers)} headers; "
          f"{len(vocab)} vocab columns.")

    groups = build_indicator_options(vocab.get("MEL Indicator", []), rows)
    print("\nIndicator options by goal:")
    for g, items in groups.items():
        print(f"  {g}: {len(items)}")
        for it in items:
            print(f"      {it['code']:12} {it['label']}")

    code2label, label2code, code2goal = flatten_options(groups)
    avail = set(code2label)

    # Simulate the Tidewater Kelp Co. example from the slides.
    profile = {
        "aqua_type": "Seaweed", "species": "Saccharina (sugar kelp)",
        "algae_type": "Brown", "farm_structure": "Longline (surface)",
        "farm_scale": "Medium (1–20 ha)", "coculture": "Monoculture",
        "climate_zone": "Temperate Atlantic",
        "site_depth": "Shallow nearshore (2–10 m)",
        "wave_energy": "Semi-exposed",
        "water_clarity": "Clear (visual methods feasible)",
    }
    sel = ["H&B 1.1.1", "WQ 1.2.1", "WQ 1.1.1", "CC 2.1.1"]
    print(f"\nFarmer-goal pre-select check (Blue Carbon): "
          f"{codes_for_farmer_goals(['Blue Carbon — ecosystem-positive farming'], avail)}")

    m = match_protocols(rows, sel, profile)
    print("\nTop match per selected indicator:")
    for code, cands in m.items():
        if cands:
            top = cands[0]
            r = top["row"]
            print(f"  {code:10} -> T{r['tier']} {r['tier_label']:12} "
                  f"[{r['skill']:12}] cost {dots(r['cost_ord'])} "
                  f"effort {dots(r['effort_ord'])} badge={top['badge']}")
            print(f"               {r['title'][:78]}")
            print(f"               method_summary present: {bool(r['method_summary'])} | "
                  f"downloads: T={bool(r['url_template'])} "
                  f"D={bool(r['url_datasheet'])} W={bool(r['url_workbook'])}")
            print(f"               ({len(cands)} candidate(s) total)")
        else:
            print(f"  {code:10} -> NO PROTOCOL IN DATABASE YET")
