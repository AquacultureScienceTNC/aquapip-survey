"""
AquaPIP — "What are you already monitoring?" screener  +  ranking boost.

Self-contained: mel_logic.py and app.py do not depend on its internals.
app.py imports this module and calls into it at three points (see INTEGRATION
at the bottom of this file).

WHAT IT DOES
------------
* Reads the `Monitoring_Approaches` tab of data/MEL_Indicator_Reference.xlsx
  BY COLUMN HEADER (same philosophy as mel_logic) — reorder or rename columns
  in Excel and nothing here breaks; add/rename/reorder rows and the screener
  updates on next app load.
* Renders the screener as one collapsible section per Category, each holding a
  multiselect of its Sub_Methods (selected count shown in the header). The
  caller supplies the page header/description.
* Scores each candidate protocol by how much of it the farmer already does, so
  the results page can float "you already do most of this" protocols up within
  their tier. It NEVER filters — a blank screener leaves order unchanged.

RANKING MODEL (see match())
---------------------------
  ag_coverage = |protocol.AG ∩ farmer.AG| / |protocol.AG|   # keyed on the
      protocol's OWN method count, so a protocol the farmer already fully
      performs scores 1.0 and one needing 3 methods they do 1 of scores 0.33.
      Rewards low marginal effort to adopt — what makes a farmer say yes.
  kw_score    = min(keyword_hits, KW_CAP) / KW_CAP          # finer signal,
      capped so a long keyword-dense summary can't run away.
  match       = AG_WEIGHT*ag_coverage + KW_WEIGHT*kw_score  # 0..1
"""

import re
import openpyxl
import streamlit as st

# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #
REF_MONITORING_SHEET = "Monitoring_Approaches"

# Ranking weights / cap (tunable).
AG_WEIGHT = 0.7
KW_WEIGHT = 0.3
KW_CAP = 3

# AG values that are meta-flags, not real methods — excluded from coverage math
# so they don't dilute the denominator.
AG_IGNORE = {"Mixed methods", "N/A — review/synthesis"}

_SESSION_KEY = "monitoring_sel"


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #
def _semis(v):
    return [x.strip() for x in str(v or "").split(";") if x.strip()]


def load(ref_path):
    """Read the Monitoring_Approaches sheet by header name.

    Returns an ordered list of categories:
        [{"category": str, "order": int,
          "submethods": [{"sub": str, "ag": set[str], "kw": list[str]}, ...]}, ...]
    Empty list if the sheet or required columns are absent (app still runs).
    """
    wb = openpyxl.load_workbook(ref_path, data_only=True, read_only=True)
    if REF_MONITORING_SHEET not in wb.sheetnames:
        return []
    ws = wb[REF_MONITORING_SHEET]

    it = ws.iter_rows(values_only=True)
    try:
        header = [(str(h).strip() if h is not None else "") for h in next(it)]
    except StopIteration:
        return []

    idx = {name: header.index(name) for name in
           ("Category_Order", "Category", "Sub_Method", "AG_Tags", "AK_Keywords")
           if name in header}
    if "Category" not in idx or "Sub_Method" not in idx:
        return []

    def cell(row, name):
        return row[idx[name]] if name in idx and idx[name] < len(row) else None

    cats, order = {}, {}
    for row in it:
        if row is None:
            continue
        cat, sub = cell(row, "Category"), cell(row, "Sub_Method")
        if not cat or not sub:
            continue
        cat, sub = str(cat).strip(), str(sub).strip()
        try:
            o = int(cell(row, "Category_Order"))
        except (TypeError, ValueError):
            o = 9999
        ag = set(_semis(cell(row, "AG_Tags")))
        kw = [k.lower() for k in _semis(cell(row, "AK_Keywords"))]
        cats.setdefault(cat, []).append({"sub": sub, "ag": ag, "kw": kw})
        order.setdefault(cat, o)

    out = [{"category": c, "order": order[c], "submethods": subs}
           for c, subs in cats.items()]
    out.sort(key=lambda d: (d["order"], d["category"]))
    return out


# --------------------------------------------------------------------------- #
# Screener page  (presentation-agnostic: caller supplies the section header)
# --------------------------------------------------------------------------- #
def render_screener(mon, key_prefix="mon"):
    """Render the screener as one collapsible section per category, each holding
    a multiselect of its sub-methods; the selected count shows in the section
    header so choices stay visible while collapsed. Writes the pooled list of
    selected sub-method dicts to st.session_state['monitoring_sel'] and returns
    it. The caller prints the page header/description. `mon` is from load()."""
    selected = []
    for cat in mon:
        c = cat["category"]
        labels = [sm["sub"] for sm in cat["submethods"]]
        by_label = {sm["sub"]: sm for sm in cat["submethods"]}
        key = f"{key_prefix}_ms_{c}"
        n = len(st.session_state.get(key, []) or [])
        header = c if not n else f"{c}   —   {n} selected"
        with st.expander(header, expanded=False):
            picks = st.multiselect(
                c, labels, key=key, label_visibility="collapsed",
                placeholder="Choose any you already do…")
            for lb in (picks or []):
                if lb in by_label:
                    selected.append(by_label[lb])
    st.session_state[_SESSION_KEY] = selected
    return selected


def get_selected():
    """Selections pooled from the screener (list of sub-method dicts)."""
    return st.session_state.get(_SESSION_KEY, [])


# --------------------------------------------------------------------------- #
# Scoring
# --------------------------------------------------------------------------- #
def _kw_hit(text_lower, kw):
    """Letter-boundary match so 'ph' does not fire inside 'phosphate' or
    'morphology', and 'tag' does not fire inside 'stage'. Digits/symbols in a
    keyword (e.g. '12s', 'gf/f', '%n', 'e. coli') are fine."""
    return re.search(r'(?<![a-z])' + re.escape(kw) + r'(?![a-z])',
                     text_lower) is not None


def _ag_set(row_ag):
    """Accept either the pre-split list (row['method_category']) or a raw
    comma-separated string; drop meta-flags."""
    if isinstance(row_ag, (list, tuple, set)):
        items = [str(x).strip() for x in row_ag]
    else:
        items = [x.strip() for x in str(row_ag or "").split(",")]
    return {x for x in items if x and x not in AG_IGNORE}


def match(row_ag, row_ak, selected,
          ag_weight=AG_WEIGHT, kw_weight=KW_WEIGHT, kw_cap=KW_CAP):
    """Score one protocol against the farmer's current monitoring.

    row_ag   : the protocol's method categories — list (row['method_category'])
               or a comma-separated string.
    row_ak   : the protocol's 'Protocol Method Summary' text.
    selected : list of sub-method dicts from render_screener()/get_selected().

    Returns (score 0..1, covered_labels list). covered_labels feeds the
    "you already do: …" chip on the result card."""
    if not selected:
        return 0.0, []

    row_ag_set = _ag_set(row_ag)
    text = str(row_ak or "").lower()

    farmer_ag = set().union(*(s["ag"] for s in selected))
    ag_cov = len(row_ag_set & farmer_ag) / len(row_ag_set) if row_ag_set else 0.0

    farmer_kw = set()
    for s in selected:
        farmer_kw.update(s["kw"])
    kw_hits = sum(1 for kw in farmer_kw if _kw_hit(text, kw))
    kw_score = min(kw_hits, kw_cap) / kw_cap

    score = ag_weight * ag_cov + kw_weight * kw_score
    covered = [s["sub"] for s in selected if s["ag"] & row_ag_set]
    return score, covered


# =========================================================================== #
# INTEGRATION  (all three edits are already applied in the app.py you were given)
# =========================================================================== #
# 0. import monitoring        # top of app.py, beside `import storage`
#    Load once (cache-keyed on the reference workbook mtime):
#        MON = monitoring.load(REF_PATH)
#
# 1. SCREENER PAGE — in render_survey(), after the farm-profile step and before
#    the indicator step:
#        step("4 - What are you already monitoring?")
#        st.caption("Optional ... only changes the order, never what you see.")
#        mon_selected = monitoring.render_screener(MON)
#    ...and carry it into the submission:  sub["mon_sel"] = mon_selected
#    Persistence: _persist_save/_persist_restore sweep "mon_" keys (mon_ms_*).
#
# 2. RANKING BOOST — in render_results(), build a scorer and pass it in:
#        sel = sub.get("mon_sel", [])
#        mon_scorer = (lambda row: monitoring.match(
#            row["method_category"], row["method_summary"], sel)) if sel else None
#        matches = ml.match_protocols(rows, sub["codes"], sub["profile"],
#                                     farmer_aqua=sub["aqua_type"],
#                                     mon_scorer=mon_scorer)
#    match_protocols adds -mon as a WITHIN-TIER sort key (after protocol/ref,
#    before farm-fit). To make fit lead instead, swap those two keys in
#    mel_logic.match_protocols.
# =========================================================================== #
