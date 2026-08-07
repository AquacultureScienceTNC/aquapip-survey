"""
TNC Filtration Service Estimator — core data + calculation logic.

Companion tool to AquaPIP. Estimates the maximum volume of water that farmed
bivalves can clear (filter), tied to MEL Objective 1.3 (farmed biomass improves
light penetration / reduces hypoxia).

Like AquaPIP's mel_logic.py, this module is deliberately Streamlit-independent
so it can be unit-tested on its own. filtration_app.py imports from here.

═══════════════════════════════════════════════════════════════════════════════
KEY DESIGN CHOICE — WHY THE TOOL UPDATES WHEN PHILINE UPDATES THE SPREADSHEET
═══════════════════════════════════════════════════════════════════════════════
Everything is read live from ONE workbook (data/Clearance_rate_estimation_tool_TNC.xlsx)
by SHEET + HEADER NAME, never by fixed position. Philine keeps editing that file
exactly the way she does today — she just adds rows/species — and the tool picks
the changes up on next load. Specifically:

  * Species list, clearance-rate (CR) equations, references, length-weight
    conversions, the "How to use" text and the caveats all come from her sheets.
  * The CR equation for each species is stored in her `Final_equations` sheet as a
    normal Excel formula that points at the calculator's input cells, e.g.
        =16.34*EXP(-0.0152*(Filtration_calculator!E2-28.6)^2)*Filtration_calculator!D2^0.608
    We AUTO-TRANSLATE that formula into a Python-evaluable expression by
    substituting the input-cell references for named variables and converting
    Excel syntax to Python:
        Filtration_calculator!E2  ->  T   (water temperature, °C)
        Filtration_calculator!C2  ->  L   (mean shell height, mm)
        Filtration_calculator!D2  ->  W   (dry tissue weight, g)
        EXP( ) -> exp( ),  ^ -> **,  strip $ absolute refs
    The result is evaluated with a SAFE evaluator (whitelisted AST — no Python
    builtins, no attribute access, no arbitrary code). If Philine ever adds a
    clean machine-readable column named exactly `Equation (py)` to Final_equations,
    that expression is used verbatim in preference to translating her Excel formula.

  * A row whose CR cell is NOT a formula (e.g. "No CR formula available",
    "CR may become available") is treated as a species that exists but is not yet
    calculable — the UI shows it greyed with the sheet's own message.

  * Which variables an equation actually uses (L vs W) is detected from the
    translated expression, so the UI can hide the dry-weight input for the one
    species (Mytilus edulis) whose equation is a function of shell height directly.
"""

import os
import re
import ast
import math
from collections import OrderedDict

import openpyxl

# --------------------------------------------------------------------------- #
# Workbook + sheet / column names (all matched by NAME, so the file stays live)
# --------------------------------------------------------------------------- #
DEFAULT_WORKBOOK = "Clearance_rate_estimation_tool_TNC.xlsx"

SHEET_HOWTO   = "How to use"
SHEET_CALC    = "Filtration_calculator"
SHEET_LW      = "length-weight_conversions"
SHEET_EQ      = "Final_equations"
SHEET_REF     = "References"

# Header names we look for inside Final_equations (case-insensitive, fuzzy).
EQ_H_SPECIES   = "species"
EQ_H_CRFORMULA = "cr formula"            # the live Excel formula that computes CR
EQ_H_INFULL    = "formula_in_full"       # human-readable version (for display)
EQ_H_REF       = "reference"
EQ_H_NOTES     = "notes"
EQ_H_PYEXPR    = "equation (py)"         # OPTIONAL clean override column

# length-weight_conversions headers
LW_H_SPECIES = "species"
LW_H_LOC     = "location"
LW_H_SEASON  = "season"
LW_H_EQTEXT  = "equation"
LW_H_EQCALC  = "length-weight conversion"   # the live formula (mm -> g DTW)
LW_H_REF     = "reference"

# References headers
REF_H_CITE = "citation"
REF_H_FULL = "full reference"
REF_H_LINK = "link"

# Variable convention used everywhere in this tool.
VAR_T = "T"   # water temperature (°C)
VAR_L = "L"   # mean shell height / length (mm)
VAR_W = "W"   # dry tissue weight (g)

# Map an input-cell COLUMN LETTER in the Filtration_calculator sheet -> variable.
# (Row number is ignored, so it is robust if Philine references C2/C3/etc.)
CALC_COL_TO_VAR = {"C": VAR_L, "D": VAR_W, "E": VAR_T}

# Phrases (in the CR-formula cell) that mean "exists but not calculable yet".
_NOT_CALCULABLE_HINTS = ("no cr", "not available", "may become", "no formula",
                         "tbc", "pending", "coming")


# --------------------------------------------------------------------------- #
# Text helpers
# --------------------------------------------------------------------------- #
def norm_species(name):
    """Normalise a species string for matching/display: kill non-breaking
    spaces, collapse whitespace, strip. (The sheets contain stray \\xa0.)"""
    if name is None:
        return ""
    s = str(name).replace("\xa0", " ").replace("\u200b", "")
    return re.sub(r"\s+", " ", s).strip()


def _norm_header(h):
    return norm_species(h).lower()


def _cell_str(v):
    if v is None:
        return ""
    return norm_species(v)


def _header_index(ws, fuzzy=True):
    """Return {normalised header -> column index (1-based)} for row 1, plus a
    resolver. Matching is case-insensitive; fuzzy=True also matches on prefix."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        h = _norm_header(ws.cell(1, c).value)
        if h:
            idx.setdefault(h, c)

    def resolve(name):
        key = _norm_header(name)
        if key in idx:
            return idx[key]
        if fuzzy:
            for h, c in idx.items():
                if h.startswith(key) or key.startswith(h):
                    return c
        return None

    return idx, resolve


# --------------------------------------------------------------------------- #
# Excel formula  ->  safe Python expression
# --------------------------------------------------------------------------- #
_CELLREF_RE = re.compile(
    r"(?:'[^']*'|" + re.escape(SHEET_CALC) + r")\s*!\s*\$?([A-Za-z]{1,3})\$?(\d+)")
_BARE_SHEET_RE = re.compile(re.escape(SHEET_CALC) + r"\s*!", re.IGNORECASE)

# Excel function name -> python (lower-cased target in our safe namespace)
_FUNC_MAP = {
    "exp": "exp", "ln": "log", "log": "log", "log10": "log10",
    "sqrt": "sqrt", "abs": "abs", "power": "pow", "pow": "pow",
    "min": "min", "max": "max", "average": "average",
}


def excel_formula_to_expr(formula):
    """Translate one Excel CR/conversion formula string into a Python expression
    using variables T, L, W. Returns the expression string (no leading '=')."""
    if formula is None:
        return ""
    s = str(formula).strip()
    if s.startswith("="):
        s = s[1:]

    # 1) input-cell references -> variables, by column letter
    def _sub_ref(m):
        col = m.group(1).upper()
        return CALC_COL_TO_VAR.get(col, f"__UNKNOWNCOL_{col}__")
    s = _CELLREF_RE.sub(_sub_ref, s)
    # any leftover "Filtration_calculator!" with odd casing
    s = _BARE_SHEET_RE.sub("", s)

    # 2) Excel operators / syntax -> Python
    s = s.replace("^", "**")
    s = s.replace("$", "")
    # normalise function names (EXP( -> exp(, LN( -> log(, POWER( -> pow( ...)
    def _sub_func(m):
        fn = m.group(1).lower()
        return _FUNC_MAP.get(fn, fn) + "("
    s = re.sub(r"([A-Za-z_][A-Za-z0-9_]*)\s*\(", _sub_func, s)

    # 3) tidy whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


# --------------------------------------------------------------------------- #
# Safe evaluator (whitelisted AST). No builtins, no names except T/L/W + funcs.
# --------------------------------------------------------------------------- #
_ALLOWED_FUNCS = {
    "exp": math.exp, "log": math.log, "log10": math.log10, "sqrt": math.sqrt,
    "abs": abs, "pow": math.pow, "min": min, "max": max,
    "average": lambda *a: sum(a) / len(a),
}
_ALLOWED_CONSTS = {"e": math.e, "pi": math.pi}
_ALLOWED_BINOPS = (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow, ast.Mod,
                   ast.FloorDiv)
_ALLOWED_UNARY = (ast.UAdd, ast.USub)


class ExprError(Exception):
    pass


def _eval_node(node, vars_):
    if isinstance(node, ast.Expression):
        return _eval_node(node.body, vars_)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float)):
            return node.value
        raise ExprError(f"disallowed constant: {node.value!r}")
    # (py<3.8 compatibility, harmless on 3.12)
    if isinstance(node, getattr(ast, "Num", ())):          # pragma: no cover
        return node.n
    if isinstance(node, ast.BinOp) and isinstance(node.op, _ALLOWED_BINOPS):
        l = _eval_node(node.left, vars_)
        r = _eval_node(node.right, vars_)
        if isinstance(node.op, ast.Add):      return l + r
        if isinstance(node.op, ast.Sub):      return l - r
        if isinstance(node.op, ast.Mult):     return l * r
        if isinstance(node.op, ast.Div):      return l / r
        if isinstance(node.op, ast.Pow):      return l ** r
        if isinstance(node.op, ast.Mod):      return l % r
        if isinstance(node.op, ast.FloorDiv): return l // r
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, _ALLOWED_UNARY):
        v = _eval_node(node.operand, vars_)
        return +v if isinstance(node.op, ast.UAdd) else -v
    if isinstance(node, ast.Name):
        if node.id in vars_:
            return vars_[node.id]
        if node.id in _ALLOWED_CONSTS:
            return _ALLOWED_CONSTS[node.id]
        raise ExprError(f"unknown name: {node.id}")
    if isinstance(node, ast.Call):
        if not isinstance(node.func, ast.Name):
            raise ExprError("only simple function calls allowed")
        fn = node.func.id.lower()
        if fn not in _ALLOWED_FUNCS:
            raise ExprError(f"function not allowed: {fn}")
        if node.keywords:
            raise ExprError("keyword args not allowed")
        args = [_eval_node(a, vars_) for a in node.args]
        return _ALLOWED_FUNCS[fn](*args)
    raise ExprError(f"disallowed expression element: {type(node).__name__}")


def safe_eval(expr, variables):
    """Evaluate a translated expression with the given variable dict. Raises
    ExprError on anything outside the whitelist."""
    if not expr:
        raise ExprError("empty expression")
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError as e:
        raise ExprError(f"syntax error: {e}")
    return _eval_node(tree, variables)


def vars_in_expr(expr):
    """Which of T/L/W actually appear in the expression (as names)."""
    used = set()
    try:
        for n in ast.walk(ast.parse(expr, mode="eval")):
            if isinstance(n, ast.Name) and n.id in (VAR_T, VAR_L, VAR_W):
                used.add(n.id)
    except SyntaxError:
        pass
    return used


# --------------------------------------------------------------------------- #
# Species / equation model
# --------------------------------------------------------------------------- #
class SpeciesEq:
    """One species' clearance-rate model."""
    def __init__(self, name, raw_cr, in_full, ref, notes, py_override=""):
        self.name = norm_species(name)
        self.raw_cr = raw_cr                       # original cell (formula or text)
        self.in_full = norm_species(in_full)       # human-readable formula
        self.reference = norm_species(ref)
        self.notes = norm_species(notes)

        raw = "" if raw_cr is None else str(raw_cr).strip()
        self._is_formula = raw.startswith("=")
        self._status_text = "" if self._is_formula else norm_species(raw_cr)

        # Prefer a clean python override column if present; else translate Excel.
        if py_override and str(py_override).strip():
            self.expr = str(py_override).strip()
        elif self._is_formula:
            self.expr = excel_formula_to_expr(raw)
        else:
            self.expr = ""

        self.uses = vars_in_expr(self.expr) if self.expr else set()

    @property
    def calculable(self):
        return bool(self.expr)

    @property
    def status_message(self):
        """Sheet's own message for a not-yet-calculable species."""
        return self._status_text or "No clearance-rate formula available yet."

    @property
    def needs_dtw(self):
        return VAR_W in self.uses

    @property
    def uses_length(self):
        return VAR_L in self.uses

    def clearance_rate(self, T, L=None, W=None):
        """Clearance rate in L per individual per hour. Returns (value, note).
        Negative results (temperature outside the fitted range) clamp to 0."""
        if not self.calculable:
            raise ExprError("species has no equation")
        v = safe_eval(self.expr, {VAR_T: float(T),
                                  VAR_L: float(L) if L is not None else float("nan"),
                                  VAR_W: float(W) if W is not None else float("nan")})
        note = ""
        if isinstance(v, complex):
            raise ExprError("complex result")
        if v != v:  # NaN
            raise ExprError("undefined result (check inputs)")
        if v < 0:
            note = ("Temperature is outside the range this species' equation was "
                    "fitted to, so the estimate was capped at 0.")
            v = 0.0
        return float(v), note


class LWConversion:
    """One shell-height -> dry-tissue-weight conversion."""
    def __init__(self, species, location, season, eq_text, eq_formula, ref):
        self.species = norm_species(species)
        self.location = norm_species(location)
        self.season = norm_species(season)
        self.eq_text = norm_species(eq_text)
        self.reference = norm_species(ref)
        self.expr = excel_formula_to_expr(eq_formula) if eq_formula else ""

    @property
    def valid(self):
        return bool(self.expr)

    def label(self):
        bits = [b for b in [self.location, self.season] if b]
        loc = " · ".join(bits) if bits else "conversion"
        return f"{loc}" + (f"  ({self.reference})" if self.reference else "")

    def dtw_from_length(self, L):
        """g dry tissue weight from shell height (mm). Clamps negatives to 0."""
        if not self.valid:
            raise ExprError("no conversion expression")
        v = safe_eval(self.expr, {VAR_L: float(L), VAR_T: float("nan"),
                                  VAR_W: float("nan")})
        if v != v:
            raise ExprError("undefined DTW")
        return max(0.0, float(v))


class Reference:
    def __init__(self, cite, full, link):
        self.cite = norm_species(cite)
        self.full = norm_species(full)
        self.link = extract_url(link)


# --------------------------------------------------------------------------- #
# HYPERLINK / URL extraction
# --------------------------------------------------------------------------- #
_HYPERLINK_RE = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)
_URL_RE = re.compile(r"https?://\S+")


def extract_url(value):
    """Get a plain URL from a cell that may hold a =HYPERLINK(...) formula, a
    bare URL, or a DOI string."""
    if value is None:
        return ""
    s = str(value).strip()
    m = _HYPERLINK_RE.search(s)
    if m:
        return m.group(1).strip()
    m = _URL_RE.search(s)
    if m:
        return m.group(0).rstrip(" ;)")
    # DOI like 10.xxxx/....  -> make it clickable
    if re.match(r"10\.\d{4,9}/\S+", s):
        return "https://doi.org/" + s
    return ""


# --------------------------------------------------------------------------- #
# Loaders
# --------------------------------------------------------------------------- #
def _get_sheet(wb, name):
    """Case-insensitive sheet lookup so minor renames don't break the tool."""
    if name in wb.sheetnames:
        return wb[name]
    low = {s.lower(): s for s in wb.sheetnames}
    key = name.lower()
    if key in low:
        return wb[low[key]]
    for s in wb.sheetnames:                      # fuzzy contains
        if key.split()[0] in s.lower():
            return wb[s]
    return None


def load_workbook_bundle(path):
    """Load everything the tool needs from the single workbook.
    Returns a dict: species (list[SpeciesEq]), conversions (list[LWConversion]),
    references (list[Reference]), howto (dict), and a species-name index."""
    wb = openpyxl.load_workbook(path, data_only=False)

    species = _load_species(wb)
    conversions = _load_conversions(wb)
    references = _load_references(wb)
    howto = _load_howto(wb)

    return {
        "species": species,
        "species_by_name": {s.name.lower(): s for s in species},
        "conversions": conversions,
        "references": references,
        "howto": howto,
    }


def _load_species(wb):
    ws = _get_sheet(wb, SHEET_EQ)
    out = []
    if ws is None:
        return out
    _, R = _header_index(ws)
    c_sp   = R(EQ_H_SPECIES)
    c_cr   = R(EQ_H_CRFORMULA)
    c_full = R(EQ_H_INFULL)
    c_ref  = R(EQ_H_REF)
    c_note = R(EQ_H_NOTES)
    c_py   = R(EQ_H_PYEXPR)
    if not c_sp:
        return out
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, c_sp).value
        if name is None or not norm_species(name):
            continue
        # A genuine species row always has something in the CR cell (a formula OR
        # a status message). Trailing "Notes" / descriptive paragraphs in col A
        # have an empty CR cell, so skip them.
        cr_val = ws.cell(r, c_cr).value if c_cr else None
        if cr_val in (None, ""):
            continue
        out.append(SpeciesEq(
            name=name,
            raw_cr=ws.cell(r, c_cr).value if c_cr else None,
            in_full=ws.cell(r, c_full).value if c_full else None,
            ref=ws.cell(r, c_ref).value if c_ref else None,
            notes=ws.cell(r, c_note).value if c_note else None,
            py_override=ws.cell(r, c_py).value if c_py else None,
        ))
    return out


def _load_conversions(wb):
    ws = _get_sheet(wb, SHEET_LW)
    out = []
    if ws is None:
        return out
    _, R = _header_index(ws)
    c_sp = R(LW_H_SPECIES); c_loc = R(LW_H_LOC); c_se = R(LW_H_SEASON)
    c_txt = R(LW_H_EQTEXT); c_calc = R(LW_H_EQCALC); c_ref = R(LW_H_REF)
    if not c_sp:
        return out
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, c_sp).value
        if name is None or not norm_species(name):
            continue
        out.append(LWConversion(
            species=name,
            location=ws.cell(r, c_loc).value if c_loc else None,
            season=ws.cell(r, c_se).value if c_se else None,
            eq_text=ws.cell(r, c_txt).value if c_txt else None,
            eq_formula=ws.cell(r, c_calc).value if c_calc else None,
            ref=ws.cell(r, c_ref).value if c_ref else None,
        ))
    return out


def _load_references(wb):
    ws = _get_sheet(wb, SHEET_REF)
    out = []
    if ws is None:
        return out
    _, R = _header_index(ws)
    c_c = R(REF_H_CITE); c_f = R(REF_H_FULL); c_l = R(REF_H_LINK)
    if not c_c:
        return out
    for r in range(2, ws.max_row + 1):
        cite = ws.cell(r, c_c).value
        if cite is None or not norm_species(cite):
            continue
        out.append(Reference(
            cite=cite,
            full=ws.cell(r, c_f).value if c_f else None,
            link=ws.cell(r, c_l).value if c_l else None,
        ))
    return out


def _load_howto(wb):
    """Parse the free-text 'How to use' sheet into intro / steps / caveats,
    live from the sheet. Robust to Philine editing the wording."""
    ws = _get_sheet(wb, SHEET_HOWTO)
    intro, steps, caveats = [], [], []
    if ws is None:
        return {"intro": intro, "steps": steps, "caveats": caveats}
    # read column A (and any adjacent) top-to-bottom
    lines = []
    for r in range(1, ws.max_row + 1):
        txt = ""
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                txt = norm_species(v)
                break
        lines.append(txt)

    section = "intro"
    for txt in lines:
        low = txt.lower()
        if not txt:
            continue
        if low in ("instructions",) or low.startswith("instruction"):
            section = "steps"; continue
        if "caveat" in low and len(txt) < 40:      # the "Important caveats" header
            section = "caveats"; continue
        if low.startswith("welcome"):
            intro.append(txt); section = "intro"; continue
        if section == "intro":
            intro.append(txt)
        elif section == "steps":
            steps.append(txt)
        else:
            caveats.append(txt)
    return {"intro": intro, "steps": steps, "caveats": caveats}


# --------------------------------------------------------------------------- #
# Reference matching (a CR-formula cell may cite several short names)
# --------------------------------------------------------------------------- #
def references_for(citation_text, all_refs):
    """Return the Reference objects whose short citation appears inside the
    species' reference string (which can list several, joined by 'and')."""
    text = norm_species(citation_text)
    if not text:
        return []
    low = text.lower()
    hits = []
    for ref in all_refs:
        if ref.cite and ref.cite.lower() in low:
            hits.append(ref)
    # if nothing matched by substring, fall back to the raw text as a pseudo-ref
    return hits


# --------------------------------------------------------------------------- #
# Per-size-class computation
# --------------------------------------------------------------------------- #
def compute_rows(sp, T, rows):
    """Compute filtration for a list of size-class rows.

    rows: list of dicts each with keys:
        shell_mm (float), count (int/float), dtw_g (float | None)
    For a length-based species, dtw_g is ignored.

    Returns (results, totals, notes):
        results: list of dicts adding cr_lph, filt_lph, filt_m3d (and echoing input)
        totals:  {'filt_lph':…, 'filt_m3d':…, 'count':…}
        notes:   set of caveat strings raised during evaluation
    """
    results = []
    tot_lph = 0.0
    tot_ct = 0.0
    notes = set()
    for rw in rows:
        L = rw.get("shell_mm")
        ct = rw.get("count") or 0
        W = rw.get("dtw_g")
        entry = {"shell_mm": L, "count": ct, "dtw_g": W,
                 "cr_lph": None, "filt_lph": None, "filt_m3d": None, "error": ""}
        try:
            if L is None or (sp.uses_length and L in (None, 0)):
                raise ExprError("enter shell height")
            if sp.needs_dtw and (W is None):
                raise ExprError("enter dry weight")
            cr, note = sp.clearance_rate(T=T, L=L, W=W)
            if note:
                notes.add(note)
            filt_lph = cr * float(ct)
            filt_m3d = filt_lph / 1000.0 * 24.0
            entry.update(cr_lph=cr, filt_lph=filt_lph, filt_m3d=filt_m3d)
            tot_lph += filt_lph
            tot_ct += float(ct)
        except ExprError as e:
            entry["error"] = str(e)
        results.append(entry)
    totals = {"filt_lph": tot_lph, "filt_m3d": tot_lph / 1000.0 * 24.0,
              "count": tot_ct}
    return results, totals, notes


# --------------------------------------------------------------------------- #
# Self-test (runs without Streamlit)
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, "data", DEFAULT_WORKBOOK)
    b = load_workbook_bundle(path)
    print(f"Loaded {len(b['species'])} species, "
          f"{len(b['conversions'])} conversion(s), "
          f"{len(b['references'])} reference(s).\n")
    for s in b["species"]:
        tag = ("calc" if s.calculable else "----")
        uses = ",".join(sorted(s.uses)) or "-"
        print(f"  [{tag}] {s.name:32s} uses={uses:5s} expr={s.expr or '('+s.status_message+')'}")
